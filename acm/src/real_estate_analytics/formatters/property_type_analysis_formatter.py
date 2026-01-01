"""
Property Type Analysis Formatter

This formatter converts property-type-specific CSV data into the market analysis format
used for comparative market analysis reports, generating both CSV and Excel outputs.
It can handle both apartment/house format and lote format.
"""

import csv
import pandas as pd
import json
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime
import logging
import os

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class PropertyTypeAnalysisFormatter:
    """Formats property-type-specific data into market analysis reports"""
    
    def __init__(self):
        """Initialize the property type analysis formatter"""
        # Define column mappings for different property types
        self.apartment_house_columns = [
            "Dirección, links", "", "VENTA", "Area Habitable", "", "Terraza", "Area Total",
            "", "administracion", "", "Alcobas", "Baños", "Parqueadero", "Walking closet",
            "LOFT", "Estudio/Sala TV", "Piso", "Alcob. Servic", "Depósito",
            "Terraza/Balcón (Área)", "Edad construcción", "Interior/Exterior", "Ascensor",
            "Calidad Acabados", "Estado, Conservación *", "Ubicación", "Estrato",
            "observaciones", "Medio de Contacto", "Contacto **"
        ]
        
        self.lote_columns = [
            "Dirección, links", "", "VENTA", "Area Util m2", "", "Precio Incluido",
            "Administración", "Area deposito", "Area Total", "Mezanine", "Baños",
            "Cocina", "Bodega", "Piso", "Parqueadero", "Año construcción", "Estrato",
            "Código inmueble", "observaciones", "Medio de Contacto", "Contacto **"
        ]
    
    def format_analysis_files(self, properties: List[Dict[str, Any]], 
                             property_address: str, 
                             output_dir: str,
                             base_filename: str = None,
                             city: str = "Bogotá",
                             property_type: str = "apartment") -> Dict[str, Any]:
        """
        Create both CSV and Excel analysis files
        
        Args:
            properties: List of property-type-specific property data dictionaries
            property_address: Address of the property being analyzed
            output_dir: Output directory for files
            base_filename: Base filename (without extension)
            city: City name (default: Bogotá)
            property_type: Type of property ('apartment', 'house', 'lote')
            
        Returns:
            Dictionary with formatting results
        """
        try:
            # Generate filenames
            if not base_filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                base_filename = f"analisis_mercado_{property_type}_{timestamp}"
            
            csv_filename = os.path.join(output_dir, f"{base_filename}.csv")
            excel_filename = os.path.join(output_dir, f"{base_filename}.xlsx")
            
            # Create output directory if it doesn't exist
            os.makedirs(output_dir, exist_ok=True)
            
            # Format to CSV
            csv_result = self.format_to_csv(properties, property_address, csv_filename, city, property_type)
            
            # Format to Excel
            excel_result = self.format_to_excel(properties, property_address, excel_filename, city, property_type)
            
            return {
                'success': csv_result['success'] and excel_result['success'],
                'csv_file': csv_filename if csv_result['success'] else None,
                'excel_file': excel_filename if excel_result['success'] else None,
                'properties_formatted': len(properties),
                'average_price': csv_result.get('average_price', 0),
                'csv_success': csv_result['success'],
                'excel_success': excel_result['success'],
                'csv_error': csv_result.get('error'),
                'excel_error': excel_result.get('error'),
                'message': f'Analysis files created: {base_filename}',
                'property_type': property_type
            }
            
        except Exception as e:
            logger.error(f"Error creating analysis files: {str(e)}")
            return {
                'success': False,
                'error': str(e),
                'csv_file': None,
                'excel_file': None,
                'properties_formatted': 0,
                'average_price': 0
            }
    
    def format_to_csv(self, properties: List[Dict[str, Any]], 
                     property_address: str, 
                     output_filename: str,
                     city: str = "Bogotá",
                     property_type: str = "apartment") -> Dict[str, Any]:
        """
        Format properties data to analysis CSV format
        
        Args:
            properties: List of property-type-specific property data dictionaries
            property_address: Address of the property being analyzed
            output_filename: Output CSV filename
            city: City name (default: Bogotá)
            property_type: Type of property ('apartment', 'house', 'lote')
            
        Returns:
            Dictionary with formatting results
        """
        try:
            # Create header
            header_rows = self.create_analysis_header(property_address, city, property_type)
            
            # Format property data
            property_rows = []
            for prop in properties:
                row = self.format_property_data(prop, property_type)
                property_rows.append(row)
            
            # Calculate average price
            average_price = self.calculate_average_price(properties, property_type)
            
            # Create footer
            footer_rows = self.create_analysis_footer(average_price, property_type)
            
            # Write to CSV
            with open(output_filename, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write header
                for row in header_rows:
                    writer.writerow(row)
                
                # Write property data
                for row in property_rows:
                    writer.writerow(row)
                
                # Write footer
                for row in footer_rows:
                    writer.writerow(row)
            
            return {
                'success': True,
                'output_file': output_filename,
                'properties_formatted': len(properties),
                'average_price': average_price,
                'property_type': property_type
            }
            
        except Exception as e:
            logger.error(f"Error formatting to CSV: {str(e)}")
            return {
                'success': False,
                'error': str(e),
                'output_file': None,
                'properties_formatted': 0,
                'average_price': 0
            }
    
    def format_to_excel(self, properties: List[Dict[str, Any]], 
                       property_address: str, 
                       output_filename: str,
                       city: str = "Bogotá",
                       property_type: str = "apartment") -> Dict[str, Any]:
        """
        Format properties data to analysis Excel format
        
        Args:
            properties: List of property-type-specific property data dictionaries
            property_address: Address of the property being analyzed
            output_filename: Output Excel filename
            city: City name (default: Bogotá)
            property_type: Type of property ('apartment', 'house', 'lote')
            
        Returns:
            Dictionary with formatting results
        """
        try:
            # Create data rows from properties
            df_data = []
            for prop in properties:
                row_data = self.format_property_data(prop, property_type)
                df_data.append(row_data)
            
            # Get column names based on property type
            if property_type.lower() == 'lote':
                columns = self.lote_columns
            else:
                columns = self.apartment_house_columns
            
            # Create Excel writer
            with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
                # Create a new workbook and worksheet
                from openpyxl import Workbook
                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = 'Análisis de Mercado'
                
                # Write header rows
                header_rows = self.create_analysis_header(property_address, city, property_type)
                current_row = 1
                for row in header_rows:
                    for col, value in enumerate(row, 1):
                        worksheet.cell(row=current_row, column=col, value=value)
                    current_row += 1
                
                # Write data rows
                for row_data in df_data:
                    for col, value in enumerate(row_data, 1):
                        worksheet.cell(row=current_row, column=col, value=value)
                    current_row += 1
                
                # Write footer rows
                footer_rows = self.create_analysis_footer(
                    self.calculate_average_price(properties, property_type), 
                    property_type
                )
                for row in footer_rows:
                    for col, value in enumerate(row, 1):
                        worksheet.cell(row=current_row, column=col, value=value)
                    current_row += 1
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            return {
                'success': True,
                'output_file': output_filename,
                'properties_formatted': len(properties),
                'property_type': property_type
            }
            
        except Exception as e:
            logger.error(f"Error formatting to Excel: {str(e)}")
            return {
                'success': False,
                'error': str(e),
                'output_file': None,
                'properties_formatted': 0
            }
    
    def format_property_data(self, property_data: Dict[str, Any], property_type: str = "apartment") -> List[str]:
        """
        Format a single property data entry into the analysis row format
        
        Args:
            property_data: Property-type-specific property data dictionary
            property_type: Type of property ('apartment', 'house', 'lote')
            
        Returns:
            List of formatted values for the analysis row
        """
        if property_type.lower() == 'lote':
            return self._format_lote_data(property_data)
        else:
            return self._format_apartment_house_data(property_data)
    
    def _format_apartment_house_data(self, property_data: Dict[str, Any]) -> List[str]:
        """Format apartment/house property data"""
        # Extract values from property-type-specific CSV format
        direccion_links = property_data.get('Dirección, links', '')
        venta = self._extract_price_value(property_data.get('VENTA', ''))
        area_habitable = self._extract_area_value(property_data.get('Area Habitable', ''))
        terraza = property_data.get('Terraza', 'NO')
        area_total = self._extract_area_value(property_data.get('Area Total', ''))
        administracion = self._extract_price_value(property_data.get('administracion', ''))
        alcobas = property_data.get('Alcobas', 0)
        banos = property_data.get('Baños', 0)
        parqueadero = property_data.get('Parqueadero', 0)
        walking_closet = property_data.get('Walking closet', 'NO')
        loft = property_data.get('LOFT', 'NO')
        estudio = property_data.get('Estudio/Sala TV', 'NO')
        piso = property_data.get('Piso', 0)
        alcob_servic = property_data.get('Alcob. Servic', 1)
        deposito = property_data.get('Depósito', 'NO')
        terraza_balcon = property_data.get('Terraza/Balcón (Área)', 'NO')
        edad_construccion = property_data.get('Edad construcción', 0)
        interior_exterior = property_data.get('Interior/Exterior', 'I')
        ascensor = property_data.get('Ascensor', 'NO')
        calidad_acabados = property_data.get('Calidad Acabados', 4)
        estado_conservacion = property_data.get('Estado, Conservación *', 4)
        ubicacion = property_data.get('Ubicación', 5)
        estrato = property_data.get('Estrato', 0)
        observaciones = property_data.get('observaciones', '')
        medio_contacto = property_data.get('Medio de Contacto', 'F.R.')
        contacto = property_data.get('Contacto **', '')
        
        # Calculate price per m2
        price_per_m2 = 0
        if area_habitable > 0 and venta > 0:
            price_per_m2 = venta / area_habitable
        
        # Format values
        return [
            direccion_links,  # Dirección, links
            "",  # Empty column
            f"{venta:,.0f}" if venta > 0 else "",  # VENTA
            f"{area_habitable:.1f}" if area_habitable > 0 else "",  # Area Habitable
            f"{price_per_m2:,.0f}" if price_per_m2 > 0 else "",  # $por m2
            terraza,  # Terraza
            f"{area_total:.1f}" if area_total > 0 else "",  # Area Total
            f"{price_per_m2:,.0f}" if price_per_m2 > 0 else "",  # $ x m2
            f"{administracion:,.0f}" if administracion > 0 else "",  # administracion
            f"{price_per_m2:,.0f}" if price_per_m2 > 0 else "",  # $ x m2
            alcobas,  # Alcobas
            banos,  # Baños
            parqueadero,  # Parqueadero
            walking_closet,  # Walking closet
            loft,  # LOFT
            estudio,  # Estudio/Sala TV
            piso,  # Piso
            alcob_servic,  # Alcob. Servic
            deposito,  # Depósito
            terraza_balcon,  # Terraza/Balcón (Área)
            edad_construccion,  # Edad construcción
            interior_exterior,  # Interior/Exterior
            ascensor,  # Ascensor
            calidad_acabados,  # Calidad Acabados
            estado_conservacion,  # Estado, Conservación *
            ubicacion,  # Ubicación
            estrato,  # Estrato
            observaciones,  # observaciones
            medio_contacto,  # Medio de Contacto
            contacto  # Contacto **
        ]
    
    def _format_lote_data(self, property_data: Dict[str, Any]) -> List[str]:
        """Format lote property data"""
        # Extract values from lote-specific CSV format
        direccion_links = property_data.get('Dirección, links', '')
        venta = self._extract_price_value(property_data.get('VENTA', ''))
        area_util = self._extract_area_value(property_data.get('Area Util m2', ''))
        precio_incluido = self._extract_price_value(property_data.get('Precio Incluido', ''))
        administracion = self._extract_price_value(property_data.get('Administración', ''))
        area_deposito = self._extract_area_value(property_data.get('Area deposito', ''))
        area_total = self._extract_area_value(property_data.get('Area Total', ''))
        mezanine = property_data.get('Mezanine', 'NO')
        banos = property_data.get('Baños', 0)
        cocina = property_data.get('Cocina', 'NO')
        bodega = property_data.get('Bodega', 'NO')
        piso = property_data.get('Piso', 0)
        parqueadero = property_data.get('Parqueadero', 0)
        ano_construccion = property_data.get('Año construcción', 0)
        estrato = property_data.get('Estrato', 0)
        codigo_inmueble = property_data.get('Código inmueble', '')
        observaciones = property_data.get('observaciones', '')
        medio_contacto = property_data.get('Medio de Contacto', 'F.R.')
        contacto = property_data.get('Contacto **', '')
        
        # Format values
        return [
            direccion_links,  # Dirección, links
            "",  # Empty column
            f"{venta:,.0f}" if venta > 0 else "",  # VENTA
            f"{area_util:.1f}" if area_util > 0 else "",  # Area Util m2
            f"{precio_incluido:,.0f}" if precio_incluido > 0 else "",  # Precio Incluido
            f"{administracion:,.0f}" if administracion > 0 else "",  # Administración
            f"{area_deposito:.1f}" if area_deposito > 0 else "",  # Area deposito
            f"{area_total:.1f}" if area_total > 0 else "",  # Area Total
            mezanine,  # Mezanine
            banos,  # Baños
            cocina,  # Cocina
            bodega,  # Bodega
            piso,  # Piso
            parqueadero,  # Parqueadero
            ano_construccion,  # Año construcción
            estrato,  # Estrato
            codigo_inmueble,  # Código inmueble
            observaciones,  # observaciones
            medio_contacto,  # Medio de Contacto
            contacto  # Contacto **
        ]
    
    def _extract_price_value(self, price_str: str) -> float:
        """Extract numeric price value from price string"""
        if not price_str:
            return 0.0
        
        # Remove common currency symbols and text
        import re
        price_clean = re.sub(r'[^\d]', '', str(price_str))
        try:
            return float(price_clean) if price_clean else 0.0
        except ValueError:
            return 0.0
    
    def _extract_area_value(self, area_str: str) -> float:
        """Extract numeric area value from area string"""
        if not area_str:
            return 0.0
        
        import re
        area_match = re.search(r'(\d+(?:\.\d+)?)', str(area_str))
        try:
            return float(area_match.group(1)) if area_match else 0.0
        except (ValueError, AttributeError):
            return 0.0
    
    def calculate_average_price(self, properties: List[Dict[str, Any]], property_type: str = "apartment") -> float:
        """Calculate average price from properties"""
        if not properties:
            return 0.0
        
        total_price = 0.0
        count = 0
        
        for prop in properties:
            if property_type.lower() == 'lote':
                price = self._extract_price_value(prop.get('VENTA', ''))
            else:
                price = self._extract_price_value(prop.get('VENTA', ''))
            
            if price > 0:
                total_price += price
                count += 1
        
        return total_price / count if count > 0 else 0.0
    
    def create_analysis_header(self, property_address: str, city: str, property_type: str) -> List[List[str]]:
        """Create analysis header rows"""
        if property_type.lower() == 'lote':
            columns = self.lote_columns
        else:
            columns = self.apartment_house_columns
        
        return [
            ["", "", "ACM-  Análisis de Mercado Comparativo", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
            ["", "", f"Fecha: {datetime.now().strftime('%B %d - %Y').upper()}", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
            ["", "", property_address, "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
            ["", "", city, "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
            ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
            columns,
            self._create_subheader_row(property_type)
        ]
    
    def _create_subheader_row(self, property_type: str) -> List[str]:
        """Create subheader row based on property type"""
        if property_type.lower() == 'lote':
            return ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""]
        else:
            return ["", "", "", "Area", "$por m2", "Area", "m2", "$ x m2", "", "$ x m2", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""]
    
    def create_analysis_footer(self, average_price: float, property_type: str) -> List[List[str]]:
        """Create analysis footer rows"""
        if property_type.lower() == 'lote':
            return [
                ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
                ["", "", f"vlr promedio", "", "", f"{average_price:,.0f}", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
                ["Elaborado por: Diana Patricia Barrera", "", f"Fecha: {datetime.now().strftime('%B %d - %Y').upper()}", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""]
            ]
        else:
            return [
                ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
                ["", "", f"vlr promedio", "", "", f"{average_price:,.0f}", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
                ["Elaborado por: Diana Patricia Barrera", "", f"Fecha: {datetime.now().strftime('%B %d - %Y').upper()}", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""]
            ]
