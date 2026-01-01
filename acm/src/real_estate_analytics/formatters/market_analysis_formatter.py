"""
Market Analysis Formatter

This formatter converts standardized property CSV data into the market analysis format
used for comparative market analysis reports, generating both CSV and Excel outputs.
"""

import csv
import pandas as pd
import json
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime
import logging
import os

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class MarketAnalysisFormatter:
    """Formats standardized property data into market analysis reports"""
    
    def __init__(self):
        """Initialize the market analysis formatter"""
        self.analysis_columns = [
            "Dirección, links",
            "",  # Empty column
            "VENTA",
            "Area Habitable",
            "",
            "Terraza",
            "Area Total",
            "",
            "administracion",
            "",
            "Alcobas",
            "Baños",
            "Parqueadero",
            "Walking closet",
            "LOFT",
            "Estudio/Sala TV",
            "Piso",
            "Alcob. Servic",
            "Depósito",
            "Terraza/Balcón (Área)",
            "Edad construcción",
            "Interior/Exterior",
            "Ascensor",
            "Calidad Acabados",
            "Estado, Conservación *",
            "Ubicación",
            "Estrato",
            "observaciones",
            "Medio de Contacto",
            "Contacto **"
        ]
        
        self.subheader_columns = [
            "",
            "",
            "",
            "Area",
            "$por m2",
            "Area",
            "m2",
            "$ x m2",
            "",
            "$ x m2",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            ""
        ]
    
    def _convert_value(self, value: Any, target_type: type = float) -> Any:
        """Convert string values to appropriate types"""
        if isinstance(value, str):
            if value == '' or value.lower() in ['none', 'null']:
                return 0 if target_type in [int, float] else ''
            try:
                if target_type == bool:
                    return value.lower() in ['true', '1', 'yes', 'si']
                elif target_type == int:
                    return int(float(value))  # Handle "85.0" -> 85
                elif target_type == float:
                    return float(value)
                else:
                    return value
            except (ValueError, TypeError):
                return 0 if target_type in [int, float] else ''
        return value

    def format_property_data(self, property_data: Dict[str, Any]) -> List[str]:
        """
        Format a single property data entry into the analysis row format
        
        Args:
            property_data: Standardized property data dictionary
            
        Returns:
            List of formatted values for the analysis row
        """
        # Extract pricing information
        pricing_raw = property_data.get('pricing', {})
        if isinstance(pricing_raw, str):
            try:
                pricing = json.loads(pricing_raw)
            except (json.JSONDecodeError, TypeError):
                pricing = {}
        else:
            pricing = pricing_raw
        area_habitable = self._convert_value(property_data.get('area_habitable', 0), float)
        area_total = self._convert_value(property_data.get('area_total', area_habitable), float)
        administration = self._convert_value(property_data.get('administration', 0), float)
        
        # Calculate price per m2
        price_per_m2 = pricing.get('price_per_m2', 0)
        total_price_per_m2 = pricing.get('total_price_per_m2', price_per_m2)
        admin_per_m2 = pricing.get('admin_per_m2', 0)
        
        # Format boolean values
        def format_boolean(value: Any) -> str:
            converted_value = self._convert_value(value, bool)
            return "SI" if converted_value else "NO"
        
        # Format numeric values
        def format_number(value: Any, decimals: int = 0) -> str:
            converted_value = self._convert_value(value, float)
            try:
                return f"{converted_value:,.{decimals}f}" if decimals > 0 else f"{int(converted_value):,}"
            except (ValueError, TypeError):
                return "0"
        
        # Build the row
        row = [
            property_data.get('address', ''),  # Dirección, links
            "",  # Empty column
            property_data.get('operation', 'VENTA'),  # VENTA
            format_number(area_habitable, 1),  # Area Habitable
            format_number(price_per_m2),  # $por m2
            format_boolean(property_data.get('terrace', False)),  # Terraza
            format_number(area_total, 1),  # Area Total
            format_number(total_price_per_m2),  # $ x m2
            format_number(administration),  # administracion
            format_number(admin_per_m2),  # $ x m2
            format_number(property_data.get('bedrooms', 0)),  # Alcobas
            format_number(property_data.get('bathrooms', 0), 1),  # Baños
            format_boolean(self._convert_value(property_data.get('parking', 0), float) > 0),  # Parqueadero
            format_boolean(property_data.get('walking_closet', False)),  # Walking closet
            format_boolean(property_data.get('loft', False)),  # LOFT
            format_boolean(property_data.get('study_room', False)),  # Estudio/Sala TV
            format_number(property_data.get('floor', 0)),  # Piso
            "0",  # Alcob. Servic (not in schema, default to 0)
            format_number(property_data.get('deposit', 0)),  # Depósito
            format_number(property_data.get('terrace_area', 0), 1),  # Terraza/Balcón (Área)
            format_number(property_data.get('construction_age', 0)),  # Edad construcción
            property_data.get('interior_exterior', 'E'),  # Interior/Exterior
            format_boolean(property_data.get('elevator', False)),  # Ascensor
            format_number(property_data.get('finish_quality', 3)),  # Calidad Acabados
            format_number(property_data.get('conservation_state', 3)),  # Estado, Conservación
            format_number(property_data.get('location_quality', 3)),  # Ubicación
            format_number(property_data.get('stratum', 3)),  # Estrato
            property_data.get('observations', ''),  # observaciones
            property_data.get('contact_method', ''),  # Medio de Contacto
            property_data.get('contact_info', '')  # Contacto
        ]
        
        return row
    
    def calculate_average_price(self, properties: List[Dict[str, Any]]) -> float:
        """
        Calculate the average price from the properties list
        
        Args:
            properties: List of property data dictionaries
            
        Returns:
            Average price as float
        """
        if not properties:
            return 0.0
        
        total_price = 0.0
        valid_prices = 0
        
        for prop in properties:
            pricing_raw = prop.get('pricing', {})
            if isinstance(pricing_raw, str):
                try:
                    pricing = json.loads(pricing_raw)
                except (json.JSONDecodeError, TypeError):
                    pricing = {}
            else:
                pricing = pricing_raw
                
            if pricing and 'price_per_m2' in pricing:
                area = self._convert_value(prop.get('area_habitable', 0), float)
                price_per_m2 = pricing.get('price_per_m2', 0)
                if area > 0 and price_per_m2 > 0:
                    total_price += area * price_per_m2
                    valid_prices += 1
        
        return total_price / valid_prices if valid_prices > 0 else 0.0
    
    def create_analysis_header(self, property_address: str, city: str = "Bogotá") -> List[List[str]]:
        """
        Create the header section for the analysis report
        
        Args:
            property_address: Address of the property being analyzed
            city: City name (default: Bogotá)
            
        Returns:
            List of header rows
        """
        current_date = datetime.now().strftime("%B %d - %Y").upper()
        
        header = [
            ["", "", "ACM-  Análisis de Mercado Comparativo", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
            ["", "", f"Fecha: {current_date}", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
            ["", "", property_address, "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
            ["", "", city, "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
            ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
            self.analysis_columns,
            self.subheader_columns
        ]
        
        return header
    
    def create_analysis_footer(self, average_price: float) -> List[List[str]]:
        """
        Create the footer section for the analysis report
        
        Args:
            average_price: Average price calculated from properties
            
        Returns:
            List of footer rows
        """
        current_date = datetime.now().strftime("%B %d - %Y").upper()
        formatted_price = f"{average_price:,.0f}"
        
        footer = [
            ["", "", "vlr promedio", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
            ["Elaborado por: Real Estate Analytics System", "", f"Fecha: {current_date}", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
            ["", "", "$  PROM  FINAL SUGERIDO INCLUIDA ADMINISTRACION", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", f'"{formatted_price}"', ""]
        ]
        
        return footer
    
    def format_to_csv(self, properties: List[Dict[str, Any]], 
                     property_address: str, 
                     output_filename: str,
                     city: str = "Bogotá") -> Dict[str, Any]:
        """
        Format properties data to analysis CSV format
        
        Args:
            properties: List of standardized property data dictionaries
            property_address: Address of the property being analyzed
            output_filename: Output CSV filename
            city: City name (default: Bogotá)
            
        Returns:
            Dictionary with formatting results
        """
        try:
            # Create header
            header_rows = self.create_analysis_header(property_address, city)
            
            # Format property data
            property_rows = []
            for prop in properties:
                row = self.format_property_data(prop)
                property_rows.append(row)
            
            # Calculate average price
            average_price = self.calculate_average_price(properties)
            
            # Create footer
            footer_rows = self.create_analysis_footer(average_price)
            
            # Write to CSV
            with open(output_filename, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write header
                for row in header_rows:
                    writer.writerow(row)
                
                # Write property data
                for row in property_rows:
                    writer.writerow(row)
                
                # Write footer
                for row in footer_rows:
                    writer.writerow(row)
            
            return {
                'success': True,
                'output_file': output_filename,
                'properties_formatted': len(property_rows),
                'average_price': average_price,
                'message': f'Analysis CSV created successfully: {output_filename}'
            }
            
        except Exception as e:
            logger.error(f"Error formatting to CSV: {str(e)}")
            return {
                'success': False,
                'error': f"Failed to format CSV: {str(e)}",
                'output_file': None
            }
    
    def format_to_excel(self, properties: List[Dict[str, Any]], 
                       property_address: str, 
                       output_filename: str,
                       city: str = "Bogotá") -> Dict[str, Any]:
        """
        Format properties data to analysis Excel format
        
        Args:
            properties: List of standardized property data dictionaries
            property_address: Address of the property being analyzed
            output_filename: Output Excel filename
            city: City name (default: Bogotá)
            
        Returns:
            Dictionary with formatting results
        """
        try:
            # Create header
            header_rows = self.create_analysis_header(property_address, city)
            
            # Format property data
            property_rows = []
            for prop in properties:
                row = self.format_property_data(prop)
                property_rows.append(row)
            
            # Calculate average price
            average_price = self.calculate_average_price(properties)
            
            # Create footer
            footer_rows = self.create_analysis_footer(average_price)
            
            # Combine all rows
            all_rows = header_rows + property_rows + footer_rows
            
            # Create DataFrame
            df = pd.DataFrame(all_rows)
            
            # Write to Excel
            with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Análisis de Mercado', index=False, header=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Análisis de Mercado']
                
                # Style the header rows
                from openpyxl.styles import Font, Alignment, PatternFill
                
                # Style the main title
                title_font = Font(bold=True, size=14)
                worksheet['C1'].font = title_font
                
                # Style the date
                date_font = Font(bold=True, size=12)
                worksheet['C2'].font = date_font
                
                # Style the address
                address_font = Font(bold=True, size=12)
                worksheet['C3'].font = address_font
                
                # Style the city
                city_font = Font(bold=True, size=12)
                worksheet['C4'].font = city_font
                
                # Style the column headers
                header_font = Font(bold=True, size=10)
                header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                for col in range(1, len(self.analysis_columns) + 1):
                    cell = worksheet.cell(row=6, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Style the subheader
                subheader_font = Font(bold=True, size=9)
                for col in range(1, len(self.subheader_columns) + 1):
                    cell = worksheet.cell(row=7, column=col)
                    cell.font = subheader_font
                
                # Style the average price row
                avg_font = Font(bold=True, size=11)
                avg_row = 6 + len(property_rows) + 1  # Header + properties + 1
                for col in range(1, len(self.analysis_columns) + 1):
                    cell = worksheet.cell(row=avg_row, column=col)
                    cell.font = avg_font
                
                # Style the final price row
                final_font = Font(bold=True, size=12, color="FF0000")  # Red color
                final_row = avg_row + 2
                for col in range(1, len(self.analysis_columns) + 1):
                    cell = worksheet.cell(row=final_row, column=col)
                    cell.font = final_font
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            return {
                'success': True,
                'output_file': output_filename,
                'properties_formatted': len(property_rows),
                'average_price': average_price,
                'message': f'Analysis Excel created successfully: {output_filename}'
            }
            
        except Exception as e:
            logger.error(f"Error formatting to Excel: {str(e)}")
            return {
                'success': False,
                'error': f"Failed to format Excel: {str(e)}",
                'output_file': None
            }
    
    def format_analysis_files(self, properties: List[Dict[str, Any]], 
                             property_address: str, 
                             output_dir: str,
                             base_filename: str = None,
                             city: str = "Bogotá") -> Dict[str, Any]:
        """
        Create both CSV and Excel analysis files
        
        Args:
            properties: List of standardized property data dictionaries
            property_address: Address of the property being analyzed
            output_dir: Output directory for files
            base_filename: Base filename (without extension)
            city: City name (default: Bogotá)
            
        Returns:
            Dictionary with formatting results
        """
        try:
            # Generate filenames
            if not base_filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                base_filename = f"analisis_mercado_{timestamp}"
            
            csv_filename = os.path.join(output_dir, f"{base_filename}.csv")
            excel_filename = os.path.join(output_dir, f"{base_filename}.xlsx")
            
            # Create output directory if it doesn't exist
            os.makedirs(output_dir, exist_ok=True)
            
            # Format to CSV
            csv_result = self.format_to_csv(properties, property_address, csv_filename, city)
            
            # Format to Excel
            excel_result = self.format_to_excel(properties, property_address, excel_filename, city)
            
            return {
                'success': csv_result['success'] and excel_result['success'],
                'csv_file': csv_filename if csv_result['success'] else None,
                'excel_file': excel_filename if excel_result['success'] else None,
                'properties_formatted': len(properties),
                'average_price': csv_result.get('average_price', 0),
                'csv_success': csv_result['success'],
                'excel_success': excel_result['success'],
                'csv_error': csv_result.get('error'),
                'excel_error': excel_result.get('error'),
                'message': f'Analysis files created: {base_filename}'
            }
            
        except Exception as e:
            logger.error(f"Error creating analysis files: {str(e)}")
            return {
                'success': False,
                'error': f"Failed to create analysis files: {str(e)}",
                'csv_file': None,
                'excel_file': None
            }


if __name__ == "__main__":
    # Test the formatter
    formatter = MarketAnalysisFormatter()
    
    # Sample property data
    sample_properties = [
        {
            'address': 'Carrera 17 No. 134-79, Cedritos',
            'operation': 'VENTA',
            'area_habitable': 85.0,
            'area_total': 85.0,
            'terrace': False,
            'administration': 350000,
            'bedrooms': 3,
            'bathrooms': 2.0,
            'parking': 1,
            'walking_closet': False,
            'loft': False,
            'study_room': False,
            'floor': 4,
            'deposit': 0,
            'terrace_area': 0,
            'construction_age': 15,
            'interior_exterior': 'E',
            'elevator': True,
            'finish_quality': 3,
            'conservation_state': 3,
            'location_quality': 3,
            'stratum': 4,
            'observations': '',
            'contact_method': 'manual_input',
            'contact_info': 'Propietario',
            'pricing': {
                'area': 85.0,
                'price_per_m2': 3294118,
                'total_area': 85.0,
                'total_price_per_m2': 3294118,
                'admin_per_m2': 0
            }
        }
    ]
    
    # Test formatting
    result = formatter.format_analysis_files(
        sample_properties,
        "Carrera 17 No. 134-79",
        "/tmp",
        "test_analysis"
    )
    
    print(f"Formatting result: {result}")
