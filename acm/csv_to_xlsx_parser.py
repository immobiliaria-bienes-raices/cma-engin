#!/usr/bin/env python3
"""
CSV to XLSX Parser - Well Formatted Excel Generator

This script converts standardized CSV files to beautifully formatted XLSX files
with proper headers, styling, colors, and formatting similar to ejemplo_venta.csv format.
"""

import pandas as pd
import csv
import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any, Optional
import argparse


class CSVToXLSXParser:
    """Converts CSV files to well-formatted XLSX files"""
    
    def __init__(self):
        """Initialize the parser with formatting styles"""
        # Define header style
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        
        # Define subheader style
        self.subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.subheader_font = Font(bold=True, size=10)
        
        # Define title style
        self.title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.title_font = Font(bold=True, color="FFFFFF", size=14)
        
        # Border style
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Alignment styles
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')
        self.right_align = Alignment(horizontal='right', vertical='center')
    
    def parse_csv_to_xlsx(self, csv_file: str, xlsx_file: str, 
                          property_address: str = None,
                          city: str = "Bogotá",
                          analyst_name: str = None) -> Dict[str, Any]:
        """
        Parse CSV file and convert to well-formatted XLSX
        
        Args:
            csv_file: Path to input CSV file
            xlsx_file: Path to output XLSX file
            property_address: Address of the property being analyzed
            city: City name
            analyst_name: Name of the analyst
            
        Returns:
            Dictionary with parsing results
        """
        try:
            # Read CSV file
            if not os.path.exists(csv_file):
                return {
                    'success': False,
                    'error': f'CSV file not found: {csv_file}'
                }
            
            # Read CSV data
            properties = []
            with open(csv_file, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                properties = list(reader)
            
            if not properties:
                return {
                    'success': False,
                    'error': 'No data found in CSV file'
                }
            
            # Extract property address from first property if not provided
            if not property_address and properties:
                property_address = properties[0].get('address', 'N/A')
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = 'Análisis de Mercado'
            
            # Write header section
            current_row = self._write_header_section(ws, property_address, city, analyst_name)
            
            # Write column headers
            current_row = self._write_column_headers(ws, current_row, properties)
            
            # Write data rows
            current_row = self._write_data_rows(ws, current_row, properties)
            
            # Write footer section
            current_row = self._write_footer_section(ws, current_row, properties)
            
            # Apply formatting
            self._apply_formatting(ws, current_row)
            
            # Auto-adjust column widths
            self._auto_adjust_column_widths(ws)
            
            # Save workbook
            wb.save(xlsx_file)
            
            return {
                'success': True,
                'input_file': csv_file,
                'output_file': xlsx_file,
                'properties_processed': len(properties),
                'total_rows': current_row
            }
            
        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()
            return {
                'success': False,
                'error': str(e),
                'error_trace': error_trace,
                'input_file': csv_file,
                'output_file': None
            }
    
    def _write_header_section(self, ws, property_address: str, city: str, analyst_name: str) -> int:
        """Write the header section with title and metadata"""
        current_row = 1
        
        # Title row
        ws.merge_cells(f'A{current_row}:AD{current_row}')
        ws[f'A{current_row}'] = 'ACM - Análisis de Mercado Comparativo'
        ws[f'A{current_row}'].fill = self.title_fill
        ws[f'A{current_row}'].font = self.title_font
        ws[f'A{current_row}'].alignment = self.center_align
        current_row += 1
        
        # Date row
        ws.merge_cells(f'A{current_row}:AD{current_row}')
        date_str = datetime.now().strftime('Fecha: %B %d - %Y').upper()
        ws[f'A{current_row}'] = date_str
        ws[f'A{current_row}'].font = Font(size=10)
        ws[f'A{current_row}'].alignment = self.left_align
        current_row += 1
        
        # Address row
        ws.merge_cells(f'A{current_row}:AD{current_row}')
        ws[f'A{current_row}'] = f' {property_address}'
        ws[f'A{current_row}'].font = Font(bold=True, size=11)
        ws[f'A{current_row}'].alignment = self.left_align
        current_row += 1
        
        # City row
        ws.merge_cells(f'A{current_row}:AD{current_row}')
        ws[f'A{current_row}'] = city
        ws[f'A{current_row}'].font = Font(size=10)
        ws[f'A{current_row}'].alignment = self.left_align
        current_row += 1
        
        # Empty row
        current_row += 1
        
        return current_row
    
    def _write_column_headers(self, ws, start_row: int, properties: List[Dict]) -> int:
        """Write column headers based on CSV structure"""
        current_row = start_row
        
        # Get column names from first property
        if properties:
            columns = list(properties[0].keys())
        else:
            return current_row
        
        # Main header row
        header_row = current_row
        col_idx = 1
        for col_name in columns:
            # Format column name for display
            display_name = col_name.replace('_', ' ').title()
            cell = ws.cell(row=header_row, column=col_idx, value=display_name)
            cell.fill = self.header_fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.thin_border
            col_idx += 1
        
        current_row += 1
        
        return current_row
    
    def _get_subheader_for_column(self, col_name: str) -> Optional[str]:
        """Get subheader text for specific columns"""
        col_lower = col_name.lower()
        if 'area' in col_lower and 'habitable' in col_lower:
            return 'Area'
        elif 'area' in col_lower and 'total' in col_lower:
            return 'm2'
        elif 'price' in col_lower or 'precio' in col_lower:
            if 'per' in col_lower or 'por' in col_lower:
                return '$por m2'
            else:
                return '$ x m2'
        elif 'administration' in col_lower or 'administracion' in col_lower:
            return '$ x m2'
        return None
    
    def _write_data_rows(self, ws, start_row: int, properties: List[Dict]) -> int:
        """Write property data rows"""
        current_row = start_row
        
        if not properties:
            return current_row
        
        columns = list(properties[0].keys())
        
        for prop in properties:
            col_idx = 1
            for col_name in columns:
                value = prop.get(col_name, '')
                
                # Format numeric values
                if isinstance(value, str):
                    # Try to convert to number if it looks like one
                    try:
                        if '.' in value:
                            value = float(value)
                        else:
                            value = int(value)
                    except:
                        pass
                
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                
                # Apply formatting based on column type
                if isinstance(value, (int, float)):
                    cell.alignment = self.right_align
                    if 'precio' in col_name.lower() or 'price' in col_name.lower():
                        cell.number_format = '#,##0'
                    elif 'area' in col_name.lower():
                        cell.number_format = '0.00'
                else:
                    cell.alignment = self.left_align
                
                cell.border = self.thin_border
                col_idx += 1
            
            current_row += 1
        
        return current_row
    
    def _write_footer_section(self, ws, start_row: int, properties: List[Dict]) -> int:
        """Write footer section with statistics"""
        current_row = start_row
        
        # Empty row
        current_row += 1
        
        # Calculate statistics
        if properties:
            # Calculate average price per m2
            prices = []
            areas = []
            for prop in properties:
                try:
                    price_per_m2 = prop.get('price_per_m2', prop.get('precio_por_m2', 0))
                    area = prop.get('area_habitable', prop.get('area_total', 0))
                    if price_per_m2 and area:
                        if isinstance(price_per_m2, str):
                            price_per_m2 = float(price_per_m2.replace(',', ''))
                        if isinstance(area, str):
                            area = float(area.replace(',', ''))
                        prices.append(price_per_m2)
                        areas.append(area)
                except:
                    pass
            
            if prices:
                avg_price = sum(prices) / len(prices)
                avg_area = sum(areas) / len(areas) if areas else 0
                total_price = avg_price * avg_area if avg_area else 0
                
                # Average price row - write values BEFORE merging
                ws[f'A{current_row}'] = 'vlr promedio'
                ws[f'A{current_row}'].font = Font(bold=True)
                ws[f'A{current_row}'].alignment = self.left_align
                
                # Average price per m2 - find column and write BEFORE any merging
                price_col = self._find_column_index(ws, 'price_per_m2') or self._find_column_index(ws, 'precio_por_m2') or 5
                if price_col:
                    price_cell = ws.cell(row=current_row, column=price_col)
                    price_cell.value = avg_price
                    price_cell.number_format = '#,##0.00'
                    price_cell.alignment = self.right_align
                
                # Now merge cells (only merge if price_col is not in the merge range)
                if price_col and price_col > 7:
                    # Only merge A-G if price is after column G
                    ws.merge_cells(f'A{current_row}:G{current_row}')
                
                current_row += 1
                
                # Analyst and final price row
                analyst_col = 1
                ws.cell(row=current_row, column=analyst_col, value='Elaborado por: [Analista]')
                ws.cell(row=current_row, column=analyst_col).font = Font(size=9)
                
                date_col = 4
                date_str = datetime.now().strftime('Fecha: %B %d - %Y').upper()
                ws.cell(row=current_row, column=date_col, value=date_str)
                ws.cell(row=current_row, column=date_col).font = Font(size=9)
                
                # Final suggested price
                final_price_col = ws.max_column - 2
                ws.cell(row=current_row, column=final_price_col, value='$ PROM FINAL SUGERIDO INCLUIDA ADMINISTRACION')
                ws.cell(row=current_row, column=final_price_col).font = Font(bold=True, size=9)
                
                final_price_col += 1
                ws.cell(row=current_row, column=final_price_col, value=total_price)
                ws.cell(row=current_row, column=final_price_col).number_format = '#,##0'
                ws.cell(row=current_row, column=final_price_col).font = Font(bold=True)
                ws.cell(row=current_row, column=final_price_col).alignment = self.right_align
                
                current_row += 1
        
        return current_row
    
    def _find_column_index(self, ws, column_name: str) -> Optional[int]:
        """Find column index by header name"""
        # Look for header row (usually row 1 or 2)
        for header_row in [1, 2, 3, 4, 5, 6]:
            for col in range(1, ws.max_column + 1):
                try:
                    cell = ws.cell(row=header_row, column=col)
                    # Check if cell is part of a merged range (skip merged cells)
                    is_merged = False
                    try:
                        for merged_range in ws.merged_cells.ranges:
                            if cell.coordinate in merged_range:
                                is_merged = True
                                break
                    except:
                        pass
                    
                    if not is_merged and hasattr(cell, 'value') and cell.value:
                        cell_value = str(cell.value)
                        if column_name.lower() in cell_value.lower():
                            return col
                except Exception as e:
                    continue
        return None
    
    def _apply_formatting(self, ws, max_row: int):
        """Apply general formatting to the worksheet"""
        # Freeze panes at header row (find first data row)
        # Skip header rows (usually rows 1-5 for title/address)
        first_data_row = 6  # Default after header section
        
        # Find where data actually starts (after column headers)
        for row in range(1, min(10, max_row)):
            try:
                cell = ws.cell(row=row, column=1)
                if hasattr(cell, 'value') and cell.value:
                    cell_str = str(cell.value).lower()
                    # Look for column header row
                    if 'address' in cell_str or 'dirección' in cell_str:
                        first_data_row = row + 1
                        break
            except:
                continue
        
        try:
            ws.freeze_panes = f'A{first_data_row}'
        except:
            pass
        
        # Apply alternating row colors (only to data rows, skip header)
        header_end_row = first_data_row - 1
        for row in range(header_end_row + 1, max_row + 1):
            if row % 2 == 0:
                fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                for col in range(1, ws.max_column + 1):
                    try:
                        cell = ws.cell(row=row, column=col)
                        # Check if cell is part of a merged range
                        is_merged = False
                        for merged_range in ws.merged_cells.ranges:
                            if cell.coordinate in merged_range:
                                is_merged = True
                                break
                        
                        if not is_merged:
                            # Only apply fill if cell doesn't already have fill
                            if not hasattr(cell, 'fill') or not cell.fill or str(cell.fill.start_color.index) == 'FFFFFFFF':
                                cell.fill = fill
                    except Exception as e:
                        # Skip cells that can't be formatted (like merged cells)
                        continue
    
    def _auto_adjust_column_widths(self, ws):
        """Auto-adjust column widths based on content"""
        for col in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            
            for row in range(1, min(ws.max_row + 1, 100)):  # Check first 100 rows
                try:
                    cell = ws.cell(row=row, column=col)
                    
                    # Check if cell is part of a merged range (skip merged cells)
                    is_merged = False
                    try:
                        for merged_range in ws.merged_cells.ranges:
                            if cell.coordinate in merged_range:
                                is_merged = True
                                break
                    except:
                        pass
                    
                    if not is_merged and hasattr(cell, 'value') and cell.value is not None:
                        try:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass
                except:
                    pass
            
            # Set column width (add some padding)
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width


def main():
    """Main function for command-line usage"""
    parser = argparse.ArgumentParser(description='Convert CSV to well-formatted XLSX')
    parser.add_argument('csv_file', help='Input CSV file path')
    parser.add_argument('-o', '--output', help='Output XLSX file path (default: same name as CSV)')
    parser.add_argument('-a', '--address', help='Property address')
    parser.add_argument('-c', '--city', default='Bogotá', help='City name (default: Bogotá)')
    parser.add_argument('-n', '--analyst', help='Analyst name')
    
    args = parser.parse_args()
    
    # Determine output file
    if args.output:
        xlsx_file = args.output
    else:
        base_name = os.path.splitext(args.csv_file)[0]
        xlsx_file = f"{base_name}.xlsx"
    
    # Create parser and convert
    parser_obj = CSVToXLSXParser()
    result = parser_obj.parse_csv_to_xlsx(
        csv_file=args.csv_file,
        xlsx_file=xlsx_file,
        property_address=args.address,
        city=args.city,
        analyst_name=args.analyst
    )
    
    if result['success']:
        print(f"✅ Successfully converted CSV to XLSX!")
        print(f"   Input:  {result['input_file']}")
        print(f"   Output: {result['output_file']}")
        print(f"   Properties processed: {result['properties_processed']}")
    else:
        print(f"❌ Error: {result.get('error', 'Unknown error')}")
        if 'error_trace' in result:
            print("\nFull error traceback:")
            print(result['error_trace'])
        sys.exit(1)


if __name__ == "__main__":
    main()

